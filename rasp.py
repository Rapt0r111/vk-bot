from datetime import date
import calendar

from openpyxl import load_workbook


# day_of_week = 'Friday'

def rasp_today(plus_day: int, group: int):
    array_messages = []
    wb = load_workbook('./Raspsianie.xlsx') if group == 1 else load_workbook('./Raspsianie1.xlsx')
    sheet = wb["Лист1"]
    day_of_week = calendar.day_name[date.today().weekday() + plus_day]
    if date.today().isocalendar().week % 2:
        week = {
            'Monday': ['B2', 'D7'],
            'Tuesday': ['B8', 'D13'],
            'Wednesday': ['B14', 'D19'],
            'Thursday': ['B20', 'D25'],
            'Friday': ['B26', 'D31'],
            'Saturday': ['B32', 'D37'],
            'Sunday': ['B69', 'F69'],
        }
    else:
        week = {
            'Monday': ['B38', 'F43'],
            'Tuesday': ['B44', 'F49'],
            'Wednesday': ['B50', 'F55'],
            'Thursday': ['B56', 'F61'],
            'Friday': ['B62', 'F67'],
            'Saturday': ['B68', 'F73'],
            'Sunday': ['B69', 'F69'],
        }

    for cellObj in sheet[week[day_of_week][0]:week[day_of_week][1]]:
        if cellObj[1].value:
            txt = ''
            for cell in cellObj:
                if cell.value:
                    txt += str(cell.value) + ' '
            array_messages.append(txt)
    return array_messages if array_messages else ['Ниче нет, спим']
